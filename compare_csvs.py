import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Set your main folder path here
main_folder = r"C:\path\to\your\folder"

# List to collect the results
results = []

# Color style for the red fill (for differences)
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Go through folders
for student_folder in os.listdir(main_folder):
    student_path = os.path.join(main_folder, student_folder)
    if os.path.isdir(student_path):
        # Get all CSV files
        csv_files = [f for f in os.listdir(student_path) if f.endswith('.csv')]
        
        if len(csv_files) == 2:
            # Read both CSV files
            file1_path = os.path.join(student_path, csv_files[0])
            file2_path = os.path.join(student_path, csv_files[1])
            
            try:
                df1 = pd.read_csv(file1_path, encoding='utf-8')
                df2 = pd.read_csv(file2_path, encoding='utf-8')
                
                # Compare shapes first
                if df1.shape != df2.shape:
                    explanation = f"Difference found: different shape {df1.shape} vs {df2.shape}"
                else:
                    # Compare the actual content
                    if df1.equals(df2):
                        explanation = "No difference"
                    else:
                        # Find where the difference is
                        diffs = (df1 != df2)
                        diff_cells = diffs.sum().sum()
                        explanation = f"Difference found: {diff_cells} different cells"

                        # Save the diff dataframe directly in the student's folder
                        if diff_cells > 0:
                            diff_file_path = os.path.join(student_path, f"{student_folder}_diffs.xlsx")
                            diffs.to_excel(diff_file_path, index=False)

                            # Apply color formatting to highlight differences (only for differences)
                            wb = load_workbook(diff_file_path)
                            ws = wb.active

                            # Loop through the dataframe and apply the formatting
                            for row in range(2, len(diffs) + 2):  # Excel rows start from 1, but pandas starts from 0
                                for col in range(1, len(diffs.columns) + 1):
                                    if diffs.iloc[row - 2, col - 1]:  # Check if there's a difference
                                        ws.cell(row=row, column=col).fill = red_fill
                                    # If there is no difference, no color fill is applied (it will remain as default/no color)

                            wb.save(diff_file_path)

            except Exception as e:
                explanation = f"Error reading files: {str(e)}"
            
            results.append({
                'Student Name': student_folder,
                'Explanation': explanation
            })
        else:
            results.append({
                'Student Name': student_folder,
                'Explanation': "Error: Expected 2 CSV files, found " + str(len(csv_files))
            })

# Save results to Excel
output_df = pd.DataFrame(results)
output_df.to_excel('comparison_result.xlsx', index=False)

print("Done! Result saved as 'comparison_result.xlsx'")
